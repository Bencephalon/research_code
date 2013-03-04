'''

Given a temporary folder and a destination folder, organizes the files downloaded form the DICOM server and creates new mask ids in the Excel spreadsheet.

Note that this script assumes the data has been collected properly. In the cases when the study is incomplete, and there are more than one scan session per maskID, upload the incomplete files manually to the directory structure!

Gustavo Sudre, 02/2013

'''

from openpyxl.reader.excel import load_workbook
import os
import glob
import tarfile
import shutil
from datetime import datetime

fname = '/Users/sudregp/MR_data/maskIds.xlsx'
tmpFolder = '/Users/sudregp/Downloads/'
mrFolder = '/Users/sudregp/MR_data/'

wb = load_workbook(filename=fname)
ws = wb.worksheets[0]

curRow = ws.get_highest_row()
curMaskId = ws.cell('A' + str(curRow)).value

dicoms = glob.glob(tmpFolder + '/*-DICOM*')
print 'Found these DICOM files in ' + tmpFolder + ':\n'
for file in dicoms:
    print file

raw_input('\nPress any key to start sorting, or Ctrl+C to quit...')

# while we still have subjects to uncompress
for file in dicoms:
    # strip out the components of the filename
    bar = file.split('-')
    subjectName = bar[0].split('/')
    subjectName = subjectName[-1]
    mrn = bar[1]
    scanId = bar[3]

    # create a new maskId and jump to the next line
    curRow += 1
    curMaskId += 1

    subjFolder = mrFolder + '/' + subjectName + '-' + mrn + '/'
    targetFolder = subjFolder + '/' + str(curMaskId)

    # if we already have the MRN in our tree, that's our target directory
    dir = os.path.dirname(targetFolder)
    # create directory if not exist
    if not os.path.exists(dir):
        os.makedirs(dir)

    # Unpack the DICOMs to the tree
    print 'Unpacking ' + file
    tar = tarfile.open(file)
    tar.extractall(targetFolder)
    tar.close()

    # Remove leading zeros from scanID
    scanId = str(int(scanId))

    rtData = glob.glob(tmpFolder + '/E' + scanId + '.*')
    if len(rtData) == 0:
        print 'WARNING: Could not find real-time data!'
        rtCopied = 'N'
    else:
        print 'Unpacking ' + rtData[0]
        tar = tarfile.open(rtData[0])
        tar.extractall(targetFolder)
        tar.close()
        rtCopied = 'Y'

    # moving extracted files inside maskId folder
    folder2move = glob.glob(targetFolder + '/*' + mrn + '/*')
    shutil.move(folder2move[0], targetFolder)
    os.rmdir(targetFolder + '/' + subjectName + '-' + mrn + '/')

    # update spreadsheet
    ws.cell('A' + str(curRow)).value = curMaskId
    ws.cell('B' + str(curRow)).value = 'Y'
    ws.cell('C' + str(curRow)).value = rtCopied

    allNames = subjectName.split('_')
    ws.cell('D' + str(curRow)).value = allNames[0]
    ws.cell('E' + str(curRow)).value = allNames[0]
    ws.cell('F' + str(curRow)).value = allNames[1]
    ws.cell('G' + str(curRow)).value = allNames[2]
    if len(allNames) > 3:
        ws.cell('H' + str(curRow)).value = allNames[3]

    ws.cell('J' + str(curRow)).value = mrn
    ws.cell('K' + str(curRow)).value = scanId

    tmp = folder2move[0].split('/')
    tmp = tmp[-1].split('-')
    myDate = datetime.strptime(tmp[0], '%Y%m%d')
    ws.cell('Q' + str(curRow)).value = myDate.strftime('%m/%d/%Y')
    ws.cell('R' + str(curRow)).value = myDate.year
    ws.cell('S' + str(curRow)).value = myDate.month
    ws.cell('T' + str(curRow)).value = myDate.day

wb.save(fname)
