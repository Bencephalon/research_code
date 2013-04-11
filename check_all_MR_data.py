''' Checks whether we have all the data listed in the spreadsheet in the server
'''

import openpyxl
import glob
from datetime import datetime


def check_for_data(dtype, date, folder):
    mask_ids = glob.glob(folder + '/*')
    for mask in mask_ids:
        dirs = glob.glob(mask + '/*')
        dir_num = [d for d, path in enumerate(dirs) if (path.find(date.strftime('%Y_%m_%d')) > 0) or (path.find(date.strftime('%Y%m%d')) > 0)]
        if len(dir_num) > 0:
            # we found a directory with the specific date inside the mask id
            # now, get the name of the text file
            txtfile = glob.glob(dirs[dir_num[0]] + '/README*.txt')
            fid = open(txtfile[0])
            text = fid.read()
            pos = text.lower().find(dtype)
            if pos > 0:
                return 1

    # if we get to here, we didn't find the type fo data in any of the matching folders
    print('Data {} not found on {} in {}').format(dtype, date.strftime('%Y_%m_%d'), folder)
    return -1


def check_date(date, folder):
    mask_ids = glob.glob(folder + '/*')
    for mask in mask_ids:
        dirs = glob.glob(mask + '/*')
        dir_num = [d for d, path in enumerate(dirs) if (path.find(scan_date.strftime('%Y_%m_%d')) > 0) or (path.find(scan_date.strftime('%Y%m%d')) > 0)]
        if len(dir_num) > 0:
            # we found a directory with the specific date inside the mask id
            return True
    return False


path = '/Volumes/neuro/MR_data/'

missing_subjects = []
has = {}
missing_mprage = []
missing_task = []
missing_rest = []
missing_edti = []

# open spreadsheet
fname = r'/Users/sudregp/Documents/Aug.2011.M.L.3T.List.xlsx'

from openpyxl.reader.excel import load_workbook
wb = load_workbook(filename=fname)
ws = wb.worksheets[0]

# get a list of the subjects in the server
subjs = glob.glob(path + '/*')

# for each subject in the spreadsheet
for row_idx in range(2, ws.get_highest_row()):
    mrn = str(ws.cell('A' + str(row_idx)).value)
    last_name = ws.cell('E' + str(row_idx)).value
    first_name = ws.cell('F' + str(row_idx)).value

    # if names have a - or space, replace by _
    last_name = last_name.replace('-', '_')
    first_name = first_name.replace('-', '_')
    last_name = last_name.replace(' ', '_')
    first_name = first_name.replace(' ', '_')

    has['rage'] = (ws.cell('W' + str(row_idx)).value == 'Y')
    has['fmri'] = (ws.cell('X' + str(row_idx)).value == 'Y')
    has['rest'] = (ws.cell('Z' + str(row_idx)).value == 'Y')
    has['edti'] = (ws.cell('AA' + str(row_idx)).value == 'Y')

    # check whether we have a folder for the subject
    dir_num = [l for l, folder in enumerate(subjs)
               if (folder.find(mrn) > 0 and
                   folder.find(last_name) > 0 and
                   folder.find(first_name) > 0)]

    # check whether the subject scanned
    scanned = ws.cell('V' + str(row_idx)).value != 'DID NOT SCAN'

    if len(dir_num) == 0 and scanned:
        # subject not found in the server. Check if we expected fMRI data
        missing = [mode for mode, val in has.iteritems() if val]
        if len(missing) > 0:
            print('Subject {}, {} ({}) not in the server. Expected: {}').format(last_name, first_name, mrn, missing)
        missing_subjects.append(mrn)
    elif scanned:
        # subject is in the server. First, check whether we have that particular scan for the subject
        scan_date = ws.cell('I' + str(row_idx)).value

        if check_date(scan_date, subjs[dir_num[0]]):
            # if we do, then look for the types of data listed in the spreadsheet
            if has_mprage:
                if check_for_data('rage', scan_date, subjs[dir_num[0]]) < 0:
                    missing_mprage.append(mrn + '_' + scan_date.strftime('%Y_%m_%d'))
            if has_task:
                if check_for_data('fmri', scan_date, subjs[dir_num[0]]) < 0:
                    missing_task.append(mrn + '_' + scan_date.strftime('%Y_%m_%d'))
            if has_rest:
                if check_for_data('rest', scan_date, subjs[dir_num[0]]) < 0:
                    missing_rest.append(mrn + '_' + scan_date.strftime('%Y_%m_%d'))
            if has_edti:
                if check_for_data('edti', scan_date, subjs[dir_num[0]]) < 0:
                    missing_edti.append(mrn + '_' + scan_date.strftime('%Y_%m_%d'))
        else:
            missing = [mode for mode, val in has.iteritems() if val]
            if len(missing) > 0:
                print('No scans on {} for {}, {} ({}). Expected: {}').format(scan_date.strftime('%Y_%m_%d'), last_name, first_name, mrn, missing)



